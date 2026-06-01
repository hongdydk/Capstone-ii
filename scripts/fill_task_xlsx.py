"""Fill data/generated/routeon_태스크양식.xlsx from fake_logistics_data_v2.csv (row per stop)."""
from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_GENERATED = ROOT / "data" / "generated"
XLSX = DATA_GENERATED / "routeon_태스크양식.xlsx"
CSV = DATA_GENERATED / "fake_logistics_data_v2.csv"
REPORT = ROOT / "scripts" / "fill_task_report.json"

HEADER_ROW = 1
DATA_START_ROW = 2

COL = {
    "task": 1,
    "type": 2,
    "place": 3,
    "address": 4,
    "recipient": 5,
    "cargo": 6,
    "tons": 7,
}


def load_csv_rows() -> list[dict[str, str]]:
    with CSV.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader)


def fill_workbook() -> dict:
    csv_rows = load_csv_rows()
    if not csv_rows:
        raise ValueError("CSV has no data rows")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["태스크"]

    if ws.max_row > HEADER_ROW:
        ws.delete_rows(DATA_START_ROW, ws.max_row - HEADER_ROW)

    row_idx = DATA_START_ROW
    task_ids: set[str] = set()

    for row in csv_rows:
        task = row.get("태스크", row.get("task_id", "")).strip()
        stop_type = row.get("구분", row.get("위치구분", "")).strip()
        place = row.get("장소명", "").strip()
        address = row.get("주소", "").strip()
        recipient = row.get("수신자", "").strip()
        cargo = row.get("화물종류", "").strip()
        tons = row.get("톤수", row.get("무게(톤)", "")).strip()

        task_ids.add(task)
        ws.cell(row_idx, COL["task"], task)
        ws.cell(row_idx, COL["type"], stop_type)
        ws.cell(row_idx, COL["place"], place)
        ws.cell(row_idx, COL["address"], address)
        if stop_type == "하차지":
            if recipient:
                ws.cell(row_idx, COL["recipient"], recipient)
            if cargo:
                ws.cell(row_idx, COL["cargo"], cargo)
            if tons:
                ws.cell(row_idx, COL["tons"], tons)
        row_idx += 1

    wb.save(XLSX)

    return {
        "sheet": "태스크",
        "header_row": HEADER_ROW,
        "data_start_row": DATA_START_ROW,
        "data_end_row": row_idx - 1,
        "csv_rows": len(csv_rows),
        "excel_data_rows": row_idx - DATA_START_ROW,
        "tasks": len(task_ids),
        "output_file": str(XLSX),
    }


if __name__ == "__main__":
    report = fill_workbook()
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {XLSX} ({report['excel_data_rows']} data rows, {report['tasks']} tasks)")
