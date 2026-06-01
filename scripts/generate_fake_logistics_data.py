"""Generate fake RouteOn task data from public logistics XLS sources.

Outputs (under data/generated/):
  - fake_logistics_stops.csv (row per stop; includes task_group, cargo_id)
  - fake_logistics_data_v2.csv (same rows, Excel-compatible columns)
  - routeon_태스크양식.xlsx (sheet 태스크, header preserved)
  - fake_logistics_orders.csv / fake_logistics_order_locations.csv (legacy flat tables)

Run from repo root: python scripts/generate_fake_logistics_data.py
"""
from __future__ import annotations

import random
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from od_stats import (
    address_to_region,
    estimate_task_distance_km,
    lookup_daily_km,
    lookup_daily_trips,
    task_variation_factor,
    ton_to_class,
)

ROOT = Path(__file__).resolve().parents[1]
DATA_SOURCE = ROOT / "data" / "source"
DATA_GENERATED = ROOT / "data" / "generated"
WAREHOUSE_XLS = DATA_SOURCE / "물류창고정보_260601.xls"
PARK_XLS = DATA_SOURCE / "물류단지정보_260601.xls"
XLSX_OUT = DATA_GENERATED / "routeon_태스크양식.xlsx"
STOPS_CSV = DATA_GENERATED / "fake_logistics_stops.csv"
V2_CSV = DATA_GENERATED / "fake_logistics_data_v2.csv"
ORDERS_CSV = DATA_GENERATED / "fake_logistics_orders.csv"
LOCATIONS_CSV = DATA_GENERATED / "fake_logistics_order_locations.csv"

NUM_TASKS = 20
RANDOM_SEED = 42

# (pickups, deliveries) — 20 tasks, mixed n:n
TASK_PATTERNS: list[tuple[int, int]] = (
    [(1, 1)] * 4
    + [(1, 2)] * 3
    + [(2, 1)] * 3
    + [(2, 2)] * 4
    + [(1, 3)] * 2
    + [(2, 3)] * 2
    + [(3, 1)] * 1
    + [(3, 2)] * 1
)

SURNAMES = ("김", "이", "박", "최", "정", "강", "조", "윤", "장", "임", "한", "오", "서", "신")
GIVEN_NAMES = (
    "민수",
    "서연",
    "지훈",
    "수빈",
    "현우",
    "예진",
    "도윤",
    "하은",
    "준호",
    "소영",
    "성민",
    "유나",
    "태양",
    "지원",
    "승현",
    "미래",
)


@dataclass(frozen=True)
class Site:
    place_name: str
    address: str
    cargo_hint: str = ""


@dataclass
class StopRow:
    task_id: int
    stop_type: str  # 상차지 | 하차지
    place_name: str
    address: str
    recipient: str = ""
    cargo_type: str = ""
    tons: str = ""
    task_group: str = ""
    cargo_id: str = ""


def _norm_addr(addr: str) -> str:
    return re.sub(r"\s+", " ", str(addr).strip())


def _clean_text(value: str, default: str = "") -> str:
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return default
    return text


def load_warehouses() -> list[Site]:
    df = pd.read_excel(WAREHOUSE_XLS, engine="xlrd")
    sites: list[Site] = []
    for _, row in df.iterrows():
        name = _clean_text(row.get("상호명", ""))
        addr = _clean_text(row.get("소재지", ""))
        cargo = _clean_text(row.get("취급품목", ""), "일반화물")
        if name and addr and name.lower() != "nan" and addr.lower() != "nan":
            sites.append(Site(name, addr, cargo))
    if not sites:
        raise RuntimeError(f"No warehouse rows in {WAREHOUSE_XLS}")
    return sites


def load_parks() -> list[Site]:
    df = pd.read_excel(PARK_XLS, engine="xlrd")
    sites: list[Site] = []
    for _, row in df.iterrows():
        name = _clean_text(row.get("물류시설명", ""))
        addr = _clean_text(row.get("주소", ""))
        if name and addr and name.lower() != "nan" and addr.lower() != "nan":
            sites.append(Site(name, addr, "종합물류"))
    return sites


def random_recipient_name(rng: random.Random) -> str:
    return f"{rng.choice(SURNAMES)}{rng.choice(GIVEN_NAMES)}"


def sample_sites(pool: list[Site], count: int, rng: random.Random) -> list[Site]:
    if count >= len(pool):
        return rng.sample(pool, len(pool))
    return rng.sample(pool, count)


def pick_delivery_site(
    pickup: Site,
    warehouses: list[Site],
    parks: list[Site],
    rng: random.Random,
    used_addresses: set[str],
) -> Site:
    """Place name and address differ from pickup when addresses differ."""
    pickup_addr = _norm_addr(pickup.address)
    candidates = [w for w in warehouses if _norm_addr(w.address) != pickup_addr]
    if parks:
        candidates.extend(parks)
    rng.shuffle(candidates)
    for site in candidates:
        addr = _norm_addr(site.address)
        if addr in used_addresses:
            continue
        if addr == pickup_addr:
            continue
        place = site.place_name
        if place == pickup.place_name:
            place = f"{site.place_name} 하차지"
        used_addresses.add(addr)
        return Site(place, site.address, site.cargo_hint or pickup.cargo_hint)

    # Fallback synthetic site
    recipient = random_recipient_name(rng)
    suffix = rng.choice(("납품지", "하차장", "입고지"))
    place = f"{recipient} {suffix}"
    addr = pickup.address  # last resort — same address => same place name allowed
    if _norm_addr(addr) != pickup_addr:
        used_addresses.add(_norm_addr(addr))
    return Site(place, addr, pickup.cargo_hint)


def build_cargo_pairs(n_pickups: int, n_deliveries: int) -> list[tuple[int, int]]:
    """Map cargo index -> (pickup_idx, delivery_idx)."""
    pairs: list[tuple[int, int]] = []
    if n_pickups == 1:
        for d in range(n_deliveries):
            pairs.append((0, d))
    elif n_deliveries == 1:
        for p in range(n_pickups):
            pairs.append((p, 0))
    else:
        n = max(n_pickups, n_deliveries)
        for i in range(n):
            pairs.append((i % n_pickups, i % n_deliveries))
    return pairs


def generate_task(
    task_id: int,
    n_pickups: int,
    n_deliveries: int,
    warehouses: list[Site],
    parks: list[Site],
    rng: random.Random,
) -> tuple[list[StopRow], dict]:
    pickups = sample_sites(warehouses, n_pickups, rng)
    used_delivery_addrs: set[str] = set()
    deliveries: list[Site] = []
    for d_idx in range(n_deliveries):
        pickup_ref = pickups[d_idx % n_pickups]
        deliveries.append(
            pick_delivery_site(pickup_ref, warehouses, parks, rng, used_delivery_addrs)
        )

    cargo_pairs = build_cargo_pairs(n_pickups, n_deliveries)
    primary_cargo = _clean_text(pickups[0].cargo_hint, "일반화물")
    total_tons = round(rng.uniform(0.8, 24.0), 1)
    per_cargo_tons = [round(total_tons / len(cargo_pairs), 1) for _ in cargo_pairs]
    per_cargo_tons[-1] = round(total_tons - sum(per_cargo_tons[:-1]), 1)

    task_group = str(task_id)
    stops: list[StopRow] = []
    shipper = pickups[0].place_name

    for i, site in enumerate(pickups):
        stops.append(
            StopRow(
                task_id=task_id,
                stop_type="상차지",
                place_name=site.place_name,
                address=site.address,
                task_group=task_group,
                cargo_id=f"T{task_id}-P{i + 1}",
            )
        )

    recipients_assigned: dict[int, str] = {}
    delivery_rows: dict[int, dict] = {}

    for cargo_idx, (p_idx, d_idx) in enumerate(cargo_pairs):
        if d_idx not in recipients_assigned:
            recipients_assigned[d_idx] = random_recipient_name(rng)
        recipient = recipients_assigned[d_idx]
        delivery = deliveries[d_idx]
        pickup = pickups[p_idx]
        cargo_type = _clean_text(pickup.cargo_hint, primary_cargo)
        if len(cargo_type) > 80:
            cargo_type = cargo_type.split(",")[0].strip()

        place = delivery.place_name
        if _norm_addr(delivery.address) != _norm_addr(pickup.address) and place == pickup.place_name:
            place = f"{recipient} 납품지"

        cargo_id = f"T{task_id}-C{cargo_idx + 1}"
        tons_val = per_cargo_tons[cargo_idx]
        key = d_idx if n_deliveries > 1 else 0
        if key not in delivery_rows:
            delivery_rows[key] = {
                "place": place,
                "address": delivery.address,
                "recipient": recipient,
                "cargo_type": cargo_type,
                "tons": tons_val,
                "cargo_ids": [cargo_id],
            }
        else:
            row = delivery_rows[key]
            row["tons"] = round(row["tons"] + tons_val, 1)
            row["cargo_ids"].append(cargo_id)
            if cargo_type not in row["cargo_type"]:
                row["cargo_type"] = f"{row['cargo_type']}, {cargo_type}".strip(", ")

    for d_idx in sorted(delivery_rows.keys()):
        row = delivery_rows[d_idx]
        stops.append(
            StopRow(
                task_id=task_id,
                stop_type="하차지",
                place_name=row["place"],
                address=row["address"],
                recipient=row["recipient"],
                cargo_type=row["cargo_type"][:120],
                tons=str(row["tons"]),
                task_group=task_group,
                cargo_id=row["cargo_ids"][0] if len(row["cargo_ids"]) == 1 else "|".join(row["cargo_ids"]),
            )
        )

    meta = {
        "order_id": task_id,
        "화주이름": shipper,
        "화물종류": primary_cargo if len(primary_cargo) <= 120 else primary_cargo[:117] + "...",
        "무게(톤)": total_tons,
        "n_pickups": n_pickups,
        "n_deliveries": n_deliveries,
    }
    return stops, meta


def _task_route_addresses(task_stops: list[StopRow]) -> list[str]:
    pickups = [s for s in task_stops if s.stop_type == "상차지"]
    deliveries = [s for s in task_stops if s.stop_type == "하차지"]
    return [s.address for s in pickups + deliveries]


def enrich_pseudonymization(stops: list[StopRow], metas: list[dict]) -> None:
    """Attach OD benchmarks and pseudonymized distances to metas and StopRow extras."""
    meta_by_id = {m["order_id"]: m for m in metas}
    by_task: dict[int, list[StopRow]] = {}
    for s in stops:
        by_task.setdefault(s.task_id, []).append(s)

    for task_id, task_stops in by_task.items():
        meta = meta_by_id[task_id]
        tons = float(meta["무게(톤)"])
        region = address_to_region(task_stops[0].address)
        tclass = ton_to_class(tons)
        route_addrs = _task_route_addresses(task_stops)
        est_km, haversine_sum, leg_kms = estimate_task_distance_km(
            task_id, region, tclass, route_addrs
        )
        factor = task_variation_factor(task_id)
        od_km = round(lookup_daily_km(region, tclass) * factor, 1)
        od_trips = round(lookup_daily_trips(region, tclass) * factor, 2)

        meta["region"] = region
        meta["ton_class"] = tclass
        meta["od_daily_km_benchmark"] = od_km
        meta["od_daily_trips_benchmark"] = od_trips
        meta["estimated_distance_km"] = est_km
        meta["haversine_sum_km"] = haversine_sum

        ordered = [s for s in task_stops if s.stop_type == "상차지"] + [
            s for s in task_stops if s.stop_type == "하차지"
        ]
        leg_idx = 0
        for i, s in enumerate(ordered):
            s._pseudonym = {
                "ton_class": tclass,
                "region": region,
                "od_daily_km_benchmark": od_km,
                "od_daily_trips_benchmark": od_trips,
                "estimated_task_distance_km": est_km,
                "haversine_leg_km": leg_kms[leg_idx] if i > 0 and leg_idx < len(leg_kms) else "",
            }
            if i > 0:
                leg_idx += 1


def write_stops_csv(path: Path, stops: list[StopRow]) -> None:
    rows = []
    for s in stops:
        row = {
            "태스크": s.task_id,
            "구분": s.stop_type,
            "장소명": s.place_name,
            "주소": s.address,
            "수신자": s.recipient,
            "화물종류": s.cargo_type,
            "톤수": s.tons,
            "task_group": s.task_group,
            "cargo_id": s.cargo_id,
        }
        pseudo = getattr(s, "_pseudonym", None)
        if pseudo:
            row.update(pseudo)
        rows.append(row)
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def write_v2_csv(path: Path, stops: list[StopRow]) -> None:
    """Excel-oriented columns (one row per stop)."""
    write_stops_csv(path, stops)


def write_legacy_csvs(stops: list[StopRow], metas: list[dict]) -> None:
    orders = pd.DataFrame(
        [
            {
                "order_id": m["order_id"],
                "화주이름": m["화주이름"],
                "화물종류": m["화물종류"],
                "무게(톤)": m["무게(톤)"],
                "task_group": str(m["order_id"]),
                "region": m.get("region", ""),
                "ton_class": m.get("ton_class", ""),
                "od_daily_km_benchmark": m.get("od_daily_km_benchmark", ""),
                "od_daily_trips_benchmark": m.get("od_daily_trips_benchmark", ""),
                "estimated_distance_km": m.get("estimated_distance_km", ""),
            }
            for m in metas
        ]
    )
    loc_rows = []
    for s in stops:
        loc_rows.append(
            {
                "order_id": s.task_id,
                "task_group": s.task_group,
                "cargo_id": s.cargo_id,
                "위치구분": s.stop_type,
                "장소명": s.place_name,
                "주소": s.address,
                "무게(톤)": s.tons or "",
                "수신자": s.recipient,
            }
        )
    orders.to_csv(ORDERS_CSV, index=False, encoding="utf-8-sig")
    pd.DataFrame(loc_rows).to_csv(LOCATIONS_CSV, index=False, encoding="utf-8-sig")


def write_xlsx(path: Path, stops: list[StopRow]) -> int:
    wb = openpyxl.load_workbook(path)
    ws = wb["태스크"]
    header_row = 1
    data_start = 2
    if ws.max_row > header_row:
        ws.delete_rows(data_start, ws.max_row - header_row)

    row_idx = data_start
    for s in stops:
        ws.cell(row_idx, 1, str(s.task_id))
        ws.cell(row_idx, 2, s.stop_type)
        ws.cell(row_idx, 3, s.place_name)
        ws.cell(row_idx, 4, s.address)
        if s.stop_type == "하차지":
            ws.cell(row_idx, 5, s.recipient)
            ws.cell(row_idx, 6, s.cargo_type)
            ws.cell(row_idx, 7, s.tons)
        row_idx += 1

    wb.save(path)
    return row_idx - data_start


def main() -> None:
    rng = random.Random(RANDOM_SEED)
    warehouses = load_warehouses()
    parks = load_parks()

    all_stops: list[StopRow] = []
    metas: list[dict] = []

    for task_id, (n_p, n_d) in enumerate(TASK_PATTERNS, start=1):
        stops, meta = generate_task(task_id, n_p, n_d, warehouses, parks, rng)
        all_stops.extend(stops)
        metas.append(meta)

    enrich_pseudonymization(all_stops, metas)

    write_stops_csv(STOPS_CSV, all_stops)
    write_v2_csv(V2_CSV, all_stops)
    write_legacy_csvs(all_stops, metas)
    excel_rows = write_xlsx(XLSX_OUT, all_stops)

    nn_tasks = [m for m in metas if m["n_pickups"] > 1 or m["n_deliveries"] > 1]
    example = next(m for m in metas if m["n_pickups"] == 2 and m["n_deliveries"] == 2)  # task 11

    print("생성 완료")
    print(f"  태스크 수: {NUM_TASKS}")
    print(f"  정류장(엑셀) 행 수: {excel_rows}")
    print(f"  n:n 태스크 수: {len(nn_tasks)}")
    print(f"  출력: {XLSX_OUT.name}, {STOPS_CSV.name}, {V2_CSV.name}")
    print(f"  예시 2상차 2하차 - 태스크 {example['order_id']}: {example['화주이름']}")


if __name__ == "__main__":
    main()
